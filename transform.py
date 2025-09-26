from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os
from dotenv import load_dotenv
import yagmail

load_dotenv()

MAIL_AUTOR = os.getenv("MAIL_AUTOR")
APP_GMAIL_PASS = os.getenv("APP_GMAIL_PASS")
MAIL_DESTINO = os.getenv("MAIL_DESTINO")

today = datetime.now()

def export_excel(data_resumen, data_sin_pa, data_informes, data_seguim, data_prest, data_alt_baj):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen general"

    headers_resumen = ["NOMBRE", "ALUMNOS", "PRESTACIONES", "INFORMES", "SEGUIMIENTOS", "PREST. CON PA", 
                       "PREST. SIN PA", "COMPLETO", "PARCIAL", "P. 2 DÍAS", "P. 3 DÍAS", "P. 4 DÍAS", 
                       "AL DÍA", "POLETTI AT"]
    ws.append(headers_resumen)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in data_resumen:
        ws.append(row)

    # Segunda hoja (Prest. sin pa por mas de 30 días)
    ws2 = wb.create_sheet(title="Sin PA por más de 30 días")

    headers_sin_pa = ["PRESTACION ID", "ALUMNO", "FEC. DE ÚLTIMA BAJA", "DÍAS SIN PA"]
    
    ws2.append(headers_sin_pa)

    for cell in ws2[1]:
      cell.font = Font(bold=True)

    for row in data_sin_pa:
        ws2.append(row)

    # Tercera hoja (informes)
    ws3 = wb.create_sheet(title="Detalle de informes")

    """ 
             COUNT(DISTINCT CASE WHEN i.informecat_nombre = "Conformidad familia" THEN i.alumnoinforme_id END) AS conf_familia,
		COUNT(DISTINCT CASE WHEN i.informecat_nombre = "Informe escolar" THEN i.alumnoinforme_id END) AS inf_escolar,
        COUNT(DISTINCT CASE WHEN i.informecat_nombre = "Informe terapéutico ext." THEN i.alumnoinforme_id END) AS inf_ter_ext,
        COUNT(DISTINCT CASE WHEN i.informecat_nombre = "Plan de trabajo - Coordinacion" THEN i.alumnoinforme_id END) AS plan_coordi """

    headers_informes = ["COORDINADORA", "ALUMNO", "DNI ALUMNO", "INF. ADMISIÓN", "INF. DIAGNÓSTICO",
                        "OTRO", "AA", "PPI", "INF. FINAL", "CONF. FLIA.", "INF. ESCOLAR", "INF. TER. EXT.",
                        "PLAN TRAB. COORD."]
    
    ws3.append(headers_informes)

    for cell in ws3[1]:
      cell.font = Font(bold=True)

    for row in data_informes:
        ws3.append(row)

    # Cuarta hoja (seguimientos)
    ws4 = wb.create_sheet(title="Detalle de seguimientos")

    headers_seguimientos = ["COORDINADORA", "PRESTACION ID", "ALUMNO", "MES", "FECHA DE CARGA", 
                            "CATEG. SEGUIMIENTO"]
    
    ws4.append(headers_seguimientos)

    for cell in ws4[1]:
      cell.font = Font(bold=True)

    for row in data_seguim:
        ws4.append(row)

    # Quinta hoja (Ultimas prest. activadas)
    ws5 = wb.create_sheet(title="Últimas prest. activadas")

    headers_prest = ["PRESTACION ID", "ALUMNO", "FECHA DE ACT."]
    
    ws5.append(headers_prest)

    for cell in ws5[1]:
      cell.font = Font(bold=True)

    for row in data_prest:
        ws5.append(row)

    # Sexta hoja (Asignaciones y bajas)
    ws6 = wb.create_sheet(title="Asignaciones y bajas")

    headers_altas_bajas = ["AÑO", "MES", "ALTAS", "BAJAS"]
    
    ws6.append(headers_altas_bajas)

    for cell in ws6[1]:
      cell.font = Font(bold=True)

    for row in data_alt_baj:
      ws6.append(row)
        
    nombre_archivo = f"reporte_coordinacion_{today.strftime('%Y-%m-%d')}.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo Excel generado: {nombre_archivo}")
    return nombre_archivo

def enviar_correo(nombre_archivo):
    try:
        yag = yagmail.SMTP(MAIL_AUTOR, APP_GMAIL_PASS)
        yag.send(
            to=MAIL_DESTINO,
            subject="Reporte general de Coordinación",
            contents= """Buenos días, se adjunta el reporte del área de Coordinación.
              \nSaludos,\nMariano López - Ailes Inclusión.""",
            attachments=nombre_archivo
        )
        print("Correo enviado correctamente.")
    except Exception as e:
        print("Error al enviar el correo:", e)




